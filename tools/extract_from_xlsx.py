#!/usr/bin/env python3
"""
One-time migration: pull James's real data out of the 2025 command-center workbook
into the JSON files the app reads.

This is NOT the yearly refresh path -- next year's refresh drops in fresh
FantasyPros / Fantasy Footballers exports (see tools/import_rankings.py).
This script exists so the rebuild starts from real data instead of placeholders,
and so the extraction is reproducible rather than a one-off paste.

Requires openpyxl. Run from the project root:
    python3 -m venv .venv && .venv/bin/pip install openpyxl
    .venv/bin/python tools/extract_from_xlsx.py
"""

import json
import os
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required:  python3 -m venv .venv && .venv/bin/pip install openpyxl")

ROOT = Path(__file__).resolve().parent.parent
SRC = ROOT / "JPC USE - Draft 2025 - in use.xlsx"
DATA = ROOT / "data"

# Data Import tab: header on row 2, players on rows 3-824.
# Column numbers are 1-indexed to match openpyxl.
COLS = {
    "pros_rank": 3,      # C  RK          - FantasyPros overall consensus rank
    "pros_tier": 4,      # D  TIERS       - FantasyPros tier
    "name": 5,           # E  PLAYER NAME
    "team": 6,           # F  TEAM
    "pos_rank_pros": 7,  # G  POS_imported - e.g. "WR1"
    "bye": 8,            # H  BYE WEEK
    "sos": 9,            # I  SOS SEASON   - e.g. "3 out of 5 stars"
    "ecr_vs_adp": 10,    # J  ECR VS. ADP
    "ffb_tier": 12,      # L  FFB_Tier
    "ffb_risk": 13,      # M  FFB_Risk
    "ffb_adp": 14,       # N  FFB_ADP      - stored as round.pick, e.g. 1.02
    "ffb_pos_rank": 15,  # O  FFB_POS_Rank
    "ffb_upside": 16,    # P  FFB_UPSIDE
    "pos": 17,           # Q  POS          - normalized, e.g. "WR"
}

# The round-by-round contingency grid on BOARD 2025 repeats in 4-column blocks.
# Each block is (taken-status col, player-name col, actual-draft-slot col).
BRANCH_BLOCKS = [
    (21, 22, 24),  # U  V  X
    (25, 26, 28),  # Y  Z  AB
    (29, 30, 32),  # AC AD AF
    (33, 34, 36),  # AG AH AJ
    (37, 38, 40),  # AK AL AN
    (41, 42, 44),  # AO AP AR   <- named "RB RB > WR WR WR"
    (45, 46, 48),  # AS AT AV   <- named "WR > RB"
    (49, 50, 52),  # AW AX AZ   <- named "RB > WR"
    (53, 54, 56),  # BA BB BD   <- named "RB > WR > TE"
]
BRANCH_HEADER_ROW = 25
BRANCH_NOTES_COL = 18  # R -- free-text round-anchored strategy notes
BRANCH_FIRST_ROW = 26  # == round 1
ROUNDS = 20  # league is 20 rounds; the 2025 board only ever got 18 filled in

DEAD = {None, "", "#N/A", "n/a", "#REF!"}


def clean(v):
    """Excel error strings and blanks all mean 'nothing here'."""
    if v in DEAD:
        return None
    if isinstance(v, str):
        v = v.strip()
        return None if v in DEAD else v
    return v


def player_name(cell):
    """Board cells hold 'Name\\nPOS - TEAM - BYE'. We only want the name."""
    v = clean(cell)
    return v.split("\n")[0].strip() if isinstance(v, str) else None


def extract_players(wb):
    ws = wb["Data Import"]
    players, seen = [], set()
    for r in range(3, ws.max_row + 1):
        name = clean(ws.cell(r, COLS["name"]).value)
        if not name:
            continue
        rec = {k: clean(ws.cell(r, c).value) for k, c in COLS.items()}
        rec["name"] = name
        # Key on name+position, not name alone: two-way players are listed once
        # per position on purpose (Travis Hunter is both a WR and a CB) and both
        # rows are real. Same name at the same position is a paste artifact.
        key = f"{name.lower()}|{(rec['pos'] or '?').lower()}"
        if key in seen:
            continue
        seen.add(key)
        rec["id"] = key
        players.append(rec)
    players.sort(key=lambda p: p["pros_rank"] if isinstance(p["pros_rank"], (int, float)) else 9999)
    return players


def extract_league(wb):
    ws = wb["BOARD 2025"]
    teams = [clean(ws.cell(1, c).value) for c in range(3, 15)]
    teams = [t for t in teams if t]
    return {
        "season": 2025,
        "teams": teams,
        # Column L on the board is James's; confirmed against the "My Team" panel.
        "myTeam": "Jimmy",
        "rounds": ROUNDS,
        "snake": True,
        # Roster requirements are NOT in either workbook -- James is confirming
        # these separately.
        #
        # This used to emit a hand-typed rosterObserved2025 dict. Every count in
        # it was wrong, and it omitted LB, DE and CB entirely -- which mattered,
        # because the missing linebackers made the 2025 draft look like it had no
        # IDP run at all. Nothing computed those numbers; someone eyeballed them.
        #
        # Left empty on purpose. The real counts are derived from the fixture in
        # data/league.json, where the board and the player pool both live, rather
        # than frozen into a migration script that cannot see either.
        "rosterObserved2025": {},
        "rosterSlots": None,
        "scoring": None,
        "keepers": None,
    }


def extract_branches(wb):
    ws = wb["BOARD 2025"]
    notes = []
    for i in range(ROUNDS):
        note = clean(ws.cell(BRANCH_FIRST_ROW + i, BRANCH_NOTES_COL).value)
        if note:
            notes.append({"round": i + 1, "text": note})

    branches = []
    for idx, (_status_col, name_col, slot_col) in enumerate(BRANCH_BLOCKS):
        label = clean(ws.cell(BRANCH_HEADER_ROW, name_col - 1).value)
        picks = []
        for i in range(ROUNDS):
            row = BRANCH_FIRST_ROW + i
            name = player_name(ws.cell(row, name_col).value)
            if not name:
                continue
            picks.append({
                "round": i + 1,
                "player": name,
                # What actually happened in 2025 -- kept as history, not as state.
                "wentAt2025": clean(ws.cell(row, slot_col).value),
            })
        if not picks:
            continue
        branches.append({
            "id": f"branch-{idx + 1}",
            "label": label or f"Plan {idx + 1}",
            "named": bool(label),
            "picks": picks,
        })
    return {"notes": notes, "branches": branches}


def extract_board_fixture(wb, league):
    """
    The 2025 draft as it actually finished, in the shape the NEW board tab uses:
    row 0 = team names, each later row = [round, pick, pick, ...] with plain names.

    This is a test fixture, not live data. It lets the app run end-to-end against
    676 real players and 216 real picks before the Google API key exists, so the
    whole thing is provably working ahead of draft day.
    """
    ws = wb["BOARD 2025"]
    rows = [[""] + league["teams"]]
    for r in range(2, 2 + ROUNDS):
        row = [r - 1]
        for c in range(3, 3 + len(league["teams"])):
            row.append(player_name(ws.cell(r, c).value) or "")
        rows.append(row)
    return rows


def main():
    if not SRC.exists():
        sys.exit(f"Source workbook not found: {SRC}")
    print(f"Reading {SRC.name} ...")
    wb = openpyxl.load_workbook(SRC, data_only=True)

    DATA.mkdir(exist_ok=True)

    players = extract_players(wb)
    (DATA / "players.2025.json").write_text(json.dumps(players, indent=2))
    print(f"  players.2025.json      {len(players)} players")

    league = extract_league(wb)
    (DATA / "league.json").write_text(json.dumps(league, indent=2))
    print(f"  league.json            {len(league['teams'])} teams, {league['rounds']} rounds")

    plan = extract_branches(wb)
    (DATA / "branches.2025.json").write_text(json.dumps(plan, indent=2))
    named = sum(1 for b in plan["branches"] if b["named"])
    print(f"  branches.2025.json     {len(plan['branches'])} branches ({named} named), "
          f"{len(plan['notes'])} round notes")

    board = extract_board_fixture(wb, league)
    (DATA / "board.local.json").write_text(json.dumps(board, indent=2))
    filled = sum(1 for row in board[1:] for cell in row[1:] if cell)
    print(f"  board.local.json       {filled} picks (2025 final, as a test fixture)")


if __name__ == "__main__":
    main()
